class Cert_gen:
    TURNO_EM_HORAS = 10
    _cert_type = None
    _worksheet = None
    _top_url = None
    _bottom_url = None
    _iter_row_min_row = 0
    _iter_row_max_col = 0
    _iter_row_max_row = 0
    _data_list = []
    _canvas = None
    _text_objects = []
    _default_font = "Helvetica"
    _bold_font = "Helvetica-Bold"
    _default_font_size = 64 - 8

    def __init__(self):
        self.define_alternative_name_list()
        self.define_cert_type(1) #Participação
        # self.define_cert_type(2) #Apresentação
        # self.define_cert_type(3) #Avaliação
        # self.define_cert_type(4) #Triagem
        self._iter_row_min_row = 90
        self._iter_row_max_col = 14
        self._iter_row_max_row = 900
        self.set_default_configs()
    
    def define_alternative_name_list(self):
        self._organizacao = [
            {"name":"Rosana", "email":""},
            {"name":"Thyéz de Oliveira Monteiro", "email":"thyezoliveira@gmail.com"},
            {"name":"Victor Di Iulio Soares", "email":"victorsaquarj@gmail.com"},
            {"name":"Bruno Corrêa dos Santos", "email":"brunocorreasantos@smec.saquarema.rj.gov.br"}
            ] # Comissão de Organização da ...
        self._avaliacao = [
            {"name":"Ilana Pereira da Costa Cunha","email":"ilana.cunha@uva.br"},
            {"name":"Flaviane Melo de Anchieta","email":"flaviane.anchieta@uva.br"},
            {"name":"Amanda Justino Acha","email":"amandaacha@unifeso.edu.br"},
            {"name":"Diogo Silva do Nascimento","email":"diogonascimento@unifeso.edu.br"},
            {"name":"Luciana Nunes Ferreira da Ponte Lopes","email":"luciananunes@unifeso.edu.br"},
            {"name":"Vanessa do Carmo Correia","email":"vanessa.correia@anhanguera.com"},
            {"name":"Wilton Araujo dos Santos","email":"wilton.santos@anhanguera.com"},
            {"name":"André Luis de Oliveira de Sant'Ana","email":"psi.andredesantanna@gmail.com"},
            {"name":"Andréa do Nascimento Sant'Anna","email":"andrea.santanna@estacio.br"},
            {"name":"Sandra Farias Miranda de Ferreira","email":"sandrafariasm@hotmail.com"},
            {"name":"Rodrigo Moura","email":"projetos.edu.marica@gmail.com"},
            {"name":"Laís Lemos Silva Novo Pinheiro","email":"lais.pinheiro@univassouras.edu.br"},
            {"name":"Camila Cristina da Silva","email":"camila.cristina@lasalle.org.br"},
            {"name":"Gessildo Mendes Júnior","email":"gessildojr@gmail.com"},
            # {"name":"Patrícia Oliveira","email":None}, # Este deve ser enviado via whatsapp
            ]

        self._triagem = []

        # APRESENTAÇÃO
        # Luciano Cesar da Costa


        #PARTICIPAÇÃO
        # Genivaldo Fábio de Souza / genivaldofabio987@gmail.com / 


    def set_default_configs(self):
        self.load_openpyxl()
        self.define_image_path()
        self.define_pdf_size()
        self.define_all_texts()
        self.iterate_worksheets(self._iter_row_min_row, self._iter_row_max_col, self._iter_row_max_row)

    def load_openpyxl(self):
        from openpyxl import load_workbook
        wb = load_workbook(filename = './signList.xlsx')
        self._worksheet = wb.active
        return self._worksheet
    
    def define_image_path(self):
        self._top_url = 'certificado_template-top.png'
        self._bottom_url = 'certificado_template-bottom.webp'

    def define_pdf_size(self):
        self._A4_landscape_custom = (3508, 2480)

    def define_all_texts(self):
        cert_string1 = ["Certificamos que", "participou da Comissão"]
        cert_string2 = ["Organizadora da","II MOSTRA DE PROJETOS E PRÁTICAS PEDAGÓGICAS INOVADORAS","da Rede Municipal de"]
        cert_string3 = "Ensino de Saquarema, nos dias 27, 28 de outubro de 2023, com carga"
        cert_string4 = ["horária de","horas, com apoio da Secretaria Municipal de Educação,"]
        cert_string5 = "Cultura, Inclusão, Ciência e Tecnologia."
        self._text = [cert_string1, cert_string2, cert_string3, cert_string4, cert_string5]

    def define_output_path(self, name):
        for data in self._data_list:
            if data["nome"] == name:
                output_path = f'pdfs/{name}.pdf'
                data["output_path"] = output_path
            
    def iterate_worksheets(self, min_row, max_col, max_row):
        # UNIDADE
        # nome = "Thyez de Oliveira Monteiro"
        # self.generate_new_data(nome, "thyezoliveiramonteiro@smec.saquarema.rj.gov.br", 40)
        # self.define_output_path(nome)

        # COMISSAO ORGANIZAÇÃO
        for row in self._organizacao:
            name = row['name']
            email = row['email']
            hours = 40
            if email != None:
                self.generate_new_data(name, email, hours)
                self.define_output_path(name)

        # COMISSAO AVALIADORA
        # for row in self._avaliacao:
        #     name = row['name']
        #     email = row['email']
        #     hours = 40
        #     if email != None:
        #         self.generate_new_data(name, email, hours)
        #         self.define_output_path(name)

        # COMISSAO TRIAGEM
        # for row in self._triagem:
        #     name = row['name']
        #     email = row['email']
        #     hours = 40
        #     if email != None:
        #         self.generate_new_data(name, email, hours)
        #         self.define_output_path(name)

        # PARTICIPAÇÃO
        # for row in self._worksheet.iter_rows(min_row=min_row, max_col=max_col, max_row=max_row):
        #     nome = None
        #     email = None
        #     multiplicador = 0
        #     horas_temp = 0
        #     if row[1].value != "1-NOME ":
        #         nome = row[1].value
        #     if row[2].value != "2-E-mail ":
        #         email = row[2].value
        #     if row[9].value != "27 - Manhã":
        #         if(row[9].value == "ok"):
        #             multiplicador += 1
        #     if row[10].value != "27 - Tarde":
        #         if(row[10].value == "ok"):
        #             multiplicador += 1
        #     if row[11].value != "28 - Manhã":
        #         if(row[11].value == "ok"):
        #             multiplicador += 1
        #     if row[12].value != "28 - Tarde":
        #         if(row[12].value == "ok"):
        #             multiplicador += 1
        #     horas_temp = self.TURNO_EM_HORAS * multiplicador
        #     if horas_temp >= 10 and email != None:
        #         self.generate_new_data(nome, email, horas_temp)
        #         self.define_output_path(nome)
    
    def generate_new_data(self, nome, email, horas_temp):
        data = {
            "nome": nome,
            "email": email,
            "horas_temp": horas_temp
        }
        self._data_list.append(data)

    def create_canvas(self, data):
        from reportlab.pdfgen import canvas
        self._canvas = canvas.Canvas(data["output_path"], pagesize=self._A4_landscape_custom)
        self.draw_images()
        self.create_default_paragraphs(data)

    def create_text_object(self, pos_x, pos_y):
        return self._canvas.beginText(pos_x, pos_y)

    def set_char_space(self, text_object, char_space):
        text_object.setCharSpace(char_space)

    def set_font(self, text_object, font_name, font_size):
        text_object.setFont(font_name, font_size)

    def set_text_line(self, text_object, text):
        text_object.textLine(text)

    def draw_text(self, text_object):
        self._canvas.drawText(text_object)
    
    def draw_images(self):
        from reportlab.lib.utils import ImageReader
        top = ImageReader(self._top_url)
        bottom = ImageReader(self._bottom_url)
        self._canvas.drawImage(top, 0, (self._A4_landscape_custom[1] - (self._A4_landscape_custom[1]/2) + 200), width=self._A4_landscape_custom[0], height=(self._A4_landscape_custom[1]/2) - 200)
        self._canvas.drawImage(bottom, 0, 0, width=self._A4_landscape_custom[0], height=(self._A4_landscape_custom[1]/3))

    def define_cert_type(self, cert_type):
        if cert_type == 1:
            print("O certificado selecionado é de participação.")
            self._cert_type = cert_type
        if cert_type == 2:
            print("O certificado selecionado é de apresentação.")
            self._cert_type = cert_type
        if cert_type == 3:
            print("O certificado selecionado é da Comissão Avaliadora.")
            self._cert_type = cert_type
        if cert_type == 4:
            print("O certificado selecionado é da Comissão de Triagem.")
            self._cert_type = cert_type
        print("Cert Type:", self._cert_type)

    def create_default_paragraphs(self, data):
        if self._cert_type == 1:
            self.create_cert_one(data)
        if self._cert_type == 2:
            self.create_cert_two(data)

    def create_cert_one(self, data):

        letter_spacing = 4
        # paragraph1_pos_x = 280 #Participação
        # paragraph2_pos_x = paragraph1_pos_x - 390 #Participação

        # paragraph1_pos_x = 180 #Avaliadora
        # paragraph2_pos_x = 190 #Avaliadora

        # paragraph1_pos_x = 180 #Triagem
        # paragraph2_pos_x = 190 #Triagem

        paragraph1_pos_x = 150 #Organização
        paragraph2_pos_x = paragraph1_pos_x - 40 #Organização


        paragraph1_1 = self.create_text_object(paragraph1_pos_x, self._A4_landscape_custom[1]/2)
        self.set_font(paragraph1_1, self._default_font, self._default_font_size)
        self.set_char_space(paragraph1_1, letter_spacing + 4)
        self.set_text_line(paragraph1_1, self._text[0][0])
        self.draw_text(paragraph1_1)

        paragraph1_2 = self.create_text_object(paragraph1_pos_x + 2425, self._A4_landscape_custom[1]/2)
        self.set_font(paragraph1_2, self._default_font, self._default_font_size)
        self.set_char_space(paragraph1_2, letter_spacing + 4)
        self.set_text_line(paragraph1_2, self._text[0][1])
        self.draw_text(paragraph1_2)


        paragraph2_1 = self.create_text_object(paragraph2_pos_x, (self._A4_landscape_custom[1]/2) - 100)
        self.set_char_space(paragraph2_1, letter_spacing)
        self.set_font(paragraph2_1, self._default_font, self._default_font_size)
        self.set_text_line(paragraph2_1, self._text[1][0])
        self.draw_text(paragraph2_1)

        paragraph2_2 = self.create_text_object(paragraph2_pos_x + 505, (self._A4_landscape_custom[1]/2) - 100)
        self.set_char_space(paragraph2_2, letter_spacing)
        self.set_font(paragraph2_2, self._bold_font, self._default_font_size)
        self.set_text_line(paragraph2_2, self._text[1][1])
        self.draw_text(paragraph2_2)

        paragraph2_3 = self.create_text_object(paragraph2_pos_x + 2625, (self._A4_landscape_custom[1]/2) - 100)
        self.set_char_space(paragraph2_3, letter_spacing)
        self.set_font(paragraph2_3, self._default_font, self._default_font_size)
        self.set_text_line(paragraph2_3, self._text[1][2])
        self.draw_text(paragraph2_3)

        paragraph3 = self.create_text_object(550, (self._A4_landscape_custom[1]/2) - 200)
        self.set_char_space(paragraph3, letter_spacing + 4)
        self.set_font(paragraph3, self._default_font, self._default_font_size)
        self.set_text_line(paragraph3, self._text[2])
        self.draw_text(paragraph3)

        paragraph4_1 = self.create_text_object(500, (self._A4_landscape_custom[1]/2) - 300)
        self.set_char_space(paragraph4_1, letter_spacing + 4)
        self.set_font(paragraph4_1, self._default_font, self._default_font_size)
        self.set_text_line(paragraph4_1, self._text[3][0])
        self.draw_text(paragraph4_1)

        paragraph4_2 = self.create_text_object(1150, (self._A4_landscape_custom[1]/2) - 300)
        self.set_char_space(paragraph4_2, letter_spacing + 4)
        self.set_font(paragraph4_2, self._default_font, self._default_font_size)
        self.set_text_line(paragraph4_2, self._text[3][1])
        self.draw_text(paragraph4_2)

        paragraph5 = self.create_text_object(1000, (self._A4_landscape_custom[1]/2) - 400)
        self.set_char_space(paragraph5, letter_spacing + 4)
        self.set_font(paragraph5, self._default_font, self._default_font_size)
        self.set_text_line(paragraph5, self._text[4])
        self.draw_text(paragraph5)

        self._canvas.line(paragraph1_pos_x + 560, (self._A4_landscape_custom[1]/2) - 10, paragraph1_pos_x + 2400, (self._A4_landscape_custom[1]/2) - 10)
        self._canvas.line(850, (self._A4_landscape_custom[1]/2) - 310, 1100, (self._A4_landscape_custom[1]/2) - 310)

        nome = self.create_text_object(paragraph1_pos_x + 580, self._A4_landscape_custom[1]/2)
        self.set_char_space(nome, 1)
        self.set_font(nome, "Courier", 58)
        self.set_text_line(nome, str(data["nome"]))
        self.draw_text(nome)
        # print("Par len: ",str(data["nome"].__len__()))

        horas_temp = self.create_text_object(950, self._A4_landscape_custom[1]/2 - 300)
        self.set_char_space(horas_temp, 1.5)
        self.set_font(horas_temp, "Courier", 58)
        self.set_text_line(horas_temp, str(data["horas_temp"]))
        self.draw_text(horas_temp)

        self.save_PDF()
        print("---------------")
        print(f"Certificado de {data['nome']} salvo em PDF com sucesso!")
        # self.send_email(data)
        print("Operação concluída com sucesso!")
        print("---------------")

    def save_PDF(self):
        self._canvas.showPage()
        self._canvas.save()
    
    def send_email(self, data):
        subjet = "Certificado de participação na II MOSTRA DE PROJETOS E PRÁTICAS PEDAGÓGICAS INOVADORAS"
        nome = data["nome"]
        msg = "Olá "+ nome +"! Segue em anexo o certificado de participação na II MOSTRA DE PROJETOS E PRÁTICAS PEDAGÓGICAS INOVADORAS da Rede Municipal de Ensino de Saquarema, nos dias 27, 28 de outubro de 2023, com carga horária máxima de 40 horas, com apoio da Secretaria Municipal de Educação, Cultura, Inclusão, Ciência e Tecnologia."
        # msg = "Olá "+ nome + "! Pedimos desculpa pelo transtorno. Reenviamos o seu certificado com os dados corrigidos!"
        email = data["email"]
        output_path = data["output_path"]
        self.enviar_email_com_anexo(email, subjet, msg, output_path, nome)
    
    def enviar_email_com_anexo(self, destinatario, assunto, mensagem, anexo_path, nome):
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.application import MIMEApplication
        from email.utils import formataddr
        
        remetente_nome = "SMECICT"
        remetente_email = 'smecict-certificados@smec.saquarema.rj.gov.br'
        senha = 'wqgc whdq ckbd hemu'
        remetente = formataddr((remetente_nome, remetente_email))
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587
        servidor = smtplib.SMTP(smtp_server, smtp_port)
        servidor.starttls()
        servidor.login(remetente_email, senha)
        mensagem_email = MIMEMultipart()
        mensagem_email['From'] = remetente
        mensagem_email['To'] = destinatario
        mensagem_email['Subject'] = assunto
        corpo_da_mensagem = mensagem
        mensagem_email.attach(MIMEText(corpo_da_mensagem, 'plain'))

        with open(anexo_path, 'rb') as anexo_arquivo:
            anexo = MIMEApplication(anexo_arquivo.read(), _subtype='pdf')
        anexo.add_header('content-disposition', 'attachment', filename=f'Certificado - {nome} - Mostra pedagogica 2023.pdf')
        mensagem_email.attach(anexo)
        servidor.sendmail(remetente_email, destinatario, mensagem_email.as_string())
        servidor.quit()

        print("Enviado para: ", destinatario)

    def clear_dir(self):
        import os, subprocess
        dir = os.getcwd() + "/pdfs"
        command = "rm -rf " + dir + "/*.pdf"
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        out, erro = process.communicate()
        if process.returncode == 0:
            print(out.decode())
        else:
            print(erro.decode())
        self.final_message()
    
    def final_message(self):
        print("=====+=====+=====+=====+=====+=====+=====+=====+=====+")
        print("Todos os certificados gerados e enviados com sucesso!")
        print("O diretorio pdf foi limpo!")
        print("=====+=====+=====+=====+=====+=====+=====+=====+=====+")
    
    def print_data(self):
        for data in self._data_list:
            self.create_canvas(data)
        # self.clear_dir()
